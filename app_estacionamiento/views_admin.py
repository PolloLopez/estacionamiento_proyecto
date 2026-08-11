# app_estacionamiento/views_admin.py
"""
Vistas del rol Admin (municipio).

Responsabilidades:
- Panel y dashboard de estadísticas
- Gestión de inspectores y vendedores
- Gestión de conductores y sus datos
- Gestión de tarifas, horarios y días especiales
- Exenciones de vehículos
- Rendiciones de caja y certificación de cierres
- Verificaciones de identidad y exenciones de conductores
- Infracciones: listado, anulación, cobro en efectivo

No incluye cobros de MercadoPago (eso es views_mp.py).
"""

from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db.models import Count, Max, Min, Q, Sum
from django.db.models.functions import TruncDate
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError as DjangoValidationError

from .decorators import require_role
from .services.infracciones import cobrar_infraccion_efectivo, MEDIOS_VALIDOS_COBRO
from .services.saldo import cargar_saldo_conductor
from .utils import sanitizar_patente
from .models import (
    CierreCaja,
    DiaEspecial,
    Estacionamiento,
    HorarioEstacionamiento,
    Infraccion,
    MovimientoCaja,
    Notificacion,
    Rendicion,
    LiquidacionComision,
    SolicitudVerificacion,
    Subcuadra,
    Tarifa,
    TIPOS_EXENCION,
    Usuario,
    Vehiculo,
    VehiculoUsuario,
)


# ─────────────────────────────────────────────────────────────────────────────
# Helper privado
# ─────────────────────────────────────────────────────────────────────────────

def _enviar_email_verificacion(correo, nombre, aprobado, motivo=""):
    """
    Envía un email al conductor informando el resultado de su verificación.
    No lanza excepciones — un email que falla no debe interrumpir el flujo.
    """
    try:
        if aprobado:
            asunto = "✅ Tu cuenta fue verificada"
            cuerpo = (
                f"Hola {nombre},\n\n"
                "¡Buenas noticias! Tu identidad fue verificada correctamente por el municipio.\n"
                "Ya podés acceder a todas las funciones de la plataforma.\n\n"
                "Sistema de Estacionamiento"
            )
        else:
            asunto = "❌ Tu verificación fue rechazada"
            cuerpo = (
                f"Hola {nombre},\n\n"
                "Tu solicitud de verificación fue rechazada."
            )
            if motivo:
                cuerpo += f"\n\nMotivo: {motivo}"
            cuerpo += (
                "\n\nPodés volver a enviar tu solicitud desde la plataforma.\n\n"
                "Sistema de Estacionamiento"
            )
        send_mail(asunto, cuerpo, None, [correo], fail_silently=True)
    except Exception:
        pass


# ─────────────────────────────────────────────────────────────────────────────
# Panel y dashboard
# ─────────────────────────────────────────────────────────────────────────────

def _error_password(password):
    """
    Valida la contraseña con dos niveles:
    1. Mínimo de 6 caracteres — siempre, en cualquier entorno.
    2. AUTH_PASSWORD_VALIDATORS de settings.py — vacíos en DEBUG=True (dev/tests),
       los 4 validadores estándar de Django en producción (DEBUG=False).
    Devuelve un string con el error, o None si la contraseña es válida.
    """
    if len(password) < 6:
        return "La contraseña debe tener al menos 6 caracteres."
    try:
        validate_password(password)
        return None
    except DjangoValidationError as e:
        return " ".join(e.messages)


@require_role("admin")
def panel_admin(request):
    """Panel principal del admin: resumen del municipio (usuarios, cobros, pendientes)."""
    usuario   = request.user
    municipio = getattr(usuario, "municipio", None)
    if not municipio:
        return redirect("login")

    infracciones_recientes = Infraccion.objects.filter(
        municipio=municipio
    ).select_related("vehiculo", "inspector").order_by("-creado_en")[:20]

    # Vehículos con estacionamiento activo en este municipio ahora mismo
    estacionamientos_activos = Estacionamiento.objects.filter(
        subcuadra__municipio=municipio,
        estado="ACTIVO",
    ).select_related("vehiculo", "subcuadra").order_by("-hora_inicio")

    verificaciones_pendientes = SolicitudVerificacion.objects.filter(
        estado="pendiente", usuario__municipio=municipio
    ).count()

    rendiciones_pendientes = CierreCaja.objects.filter(
        usuario__municipio=municipio, certificado=False,
    ).count()

    from django.urls import reverse as _reverse

    # Ítems del sidebar de gestión: label, url, badge (opcional)
    sidebar_gestion = [
        {"label": "👤 Usuarios",         "url": _reverse("gestionar_usuarios"),       "badge": None},
        {"label": "👮 Inspectores",       "url": _reverse("gestionar_inspectores"),    "badge": None},
        {"label": "💰 Vendedores",        "url": _reverse("gestionar_vendedores"),     "badge": None},
        {"label": "🚗 Vehículos",         "url": _reverse("admin_vehiculos"),          "badge": None},
        {"label": "📋 Infracciones",      "url": _reverse("admin_infracciones"),       "badge": None},
        {"label": "🚫 Exenciones",        "url": _reverse("exenciones"),               "badge": None},
        {"label": "📍 Subcuadras GPS",    "url": _reverse("gestionar_subcuadras"),     "badge": None},
        {"label": "💲 Tarifas",           "url": _reverse("gestionar_tarifas"),        "badge": None},
        {"label": "🕐 Horarios",          "url": _reverse("gestionar_horarios"),       "badge": None},
        {"label": "📅 Días especiales",   "url": _reverse("gestionar_dias_especiales"),"badge": None},
        {"label": "✅ Verificaciones",    "url": _reverse("gestionar_verificaciones"), "badge": verificaciones_pendientes or None},
        {"label": "💼 Rendiciones",       "url": _reverse("admin_rendiciones"),        "badge": rendiciones_pendientes or None},
    ]

    return render(request, "admin/panel_admin.html", {
        "infracciones_recientes":    infracciones_recientes,
        "estacionamientos_activos":  estacionamientos_activos,
        "verificaciones_pendientes": verificaciones_pendientes,
        "rendiciones_pendientes":    rendiciones_pendientes,
        "sidebar_gestion":           sidebar_gestion,
    })


@require_role("admin")
def dashboard_admin(request):
    """Dashboard de estadísticas: infracciones por inspector, patentes por día, cobros."""
    municipio = request.user.municipio

    infracciones_por_inspector = Infraccion.objects.filter(
        municipio=municipio
    ).values("inspector__correo").annotate(total=Count("id")).order_by("-total")

    patentes_por_dia = Vehiculo.objects.filter(
        municipio=municipio
    ).annotate(fecha=TruncDate("fecha_creacion")).values("fecha").annotate(total=Count("id"))

    cobros = MovimientoCaja.objects.filter(
        usuario__municipio=municipio
    ).values("usuario__correo").annotate(total=Sum("monto")).order_by("-total")

    return render(request, "admin/panel_admin.html", {
        "infracciones_por_inspector": infracciones_por_inspector,
        "patentes_por_dia":           patentes_por_dia,
        "cobros":                     cobros,
    })


@require_role("admin")
def inicio_admin(request):
    """Alias de entrada al admin — redirige al panel principal."""
    return redirect("panel_admin")


# ─────────────────────────────────────────────────────────────────────────────
# Exenciones de vehículos
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def panel_exenciones(request):
    """
    Busca un vehículo por patente y gestiona su exención (global o parcial por subcuadra).
    Puede recibir ?patente=XYZ desde detalle_usuario para pre-cargar.
    """
    usuario   = request.user
    municipio = getattr(usuario, "municipio", None)
    if not municipio:
        return redirect("login")

    subcuadras = Subcuadra.objects.filter(municipio=municipio)
    vehiculo   = None
    accion     = request.POST.get("accion")

    def _buscar_vehiculo(patente):
        return Vehiculo.objects.filter(patente=patente).filter(
            Q(municipio=municipio) | Q(municipio__isnull=True)
        ).first()

    # Pre-carga desde detalle_usuario con ?patente=
    patente_get = sanitizar_patente(request.GET.get("patente", ""))
    if patente_get and not accion:
        vehiculo = _buscar_vehiculo(patente_get)

    if request.method == "POST":
        if accion == "buscar":
            patente  = sanitizar_patente(request.POST.get("patente") or "")
            vehiculo = _buscar_vehiculo(patente)

        elif accion == "crear_vehiculo":
            # El admin confirma crear un vehículo que no existe
            patente = sanitizar_patente(request.POST.get("patente") or "")
            tipo    = request.POST.get("tipo", "auto")
            if tipo not in ("auto", "moto"):
                tipo = "auto"
            if patente:
                vehiculo, _ = Vehiculo.objects.get_or_create(
                    patente=patente,
                    defaults={"tipo": tipo, "municipio": municipio},
                )
                messages.success(request, f"Vehículo {patente} creado. Configurá su exención.")

        elif accion == "guardar":
            patente  = sanitizar_patente(request.POST.get("patente") or "")
            vehiculo = _buscar_vehiculo(patente)

            if vehiculo:
                vehiculo.exento_global  = request.POST.get("exento_global") == "on"
                vehiculo.tipo_exencion  = request.POST.get("tipo_exencion") or None
                vehiculo.notas_exencion = request.POST.get("notas_exencion", "").strip() or None
                vehiculo.save()

                subcuadras_ids    = request.POST.getlist("subcuadras")
                subcuadras_validas = Subcuadra.objects.filter(
                    id__in=subcuadras_ids, municipio=municipio
                ).values_list("id", flat=True)
                vehiculo.subcuadras_exentas.set(subcuadras_validas)

                messages.success(request, f"✅ Exención guardada para {vehiculo.patente}.")
            else:
                messages.error(request, "No se encontró el vehículo con esa patente.")

        elif accion == "verificar":
            # El admin confirma que contactó al titular y completó los datos.
            # Marca el vehículo como verificado.
            patente  = sanitizar_patente(request.POST.get("patente") or "")
            vehiculo = _buscar_vehiculo(patente)
            if vehiculo:
                vehiculo.exencion_verificada = True
                vehiculo.save()
                messages.success(request, f"✅ Exención de {patente} marcada como verificada.")
            else:
                messages.error(request, "No se encontró el vehículo.")

    # Listado global: todos los vehículos con alguna exención en el municipio
    qs_exentos = Vehiculo.objects.filter(
        Q(exento_global=True, municipio=municipio)
        | Q(subcuadras_exentas__municipio=municipio)
    ).distinct().prefetch_related(
        "vehiculousuario_set__usuario", "subcuadras_exentas"
    )

    # Pendientes de verificación: importados que todavía no se verificaron
    vehiculos_pendientes = qs_exentos.filter(exencion_verificada=False).order_by("patente")
    vehiculos_exentos    = qs_exentos.filter(exencion_verificada=True).order_by("patente")

    return render(request, "admin/exenciones.html", {
        "vehiculo":             vehiculo,
        "subcuadras":           subcuadras,
        "tipos_exencion":       TIPOS_EXENCION,
        "vehiculos_exentos":    vehiculos_exentos,
        "vehiculos_pendientes": vehiculos_pendientes,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Saldo de conductores
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def cargar_saldo(request, usuario_id):
    """El admin carga saldo a un conductor. Registra el movimiento en la caja del admin."""
    admin   = request.user
    usuario = get_object_or_404(Usuario, id=usuario_id, municipio=admin.municipio)

    comprobante = None

    if request.method == "POST":
        monto_str = request.POST.get("monto", "")
        try:
            monto = Decimal(monto_str)
            cargar_saldo_conductor(admin=admin, conductor=usuario, monto=monto)
            # Refresca el usuario para obtener el saldo actualizado
            usuario.refresh_from_db()
            comprobante = {
                "monto":      monto,
                "saldo_nuevo": usuario.saldo,
                "fecha":      timezone.localtime(),
                "admin":      admin,
            }
        except (ValueError, Exception):
            return render(request, "admin/cargar_saldo.html", {
                "usuario": usuario,
                "error": "Monto inválido",
            })

    return render(request, "admin/cargar_saldo.html", {
        "usuario":     usuario,
        "comprobante": comprobante,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Gestión de inspectores
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def gestionar_inspectores(request):
    """Lista, crea y configura inspectores del municipio."""
    usuario   = request.user
    municipio = usuario.municipio
    error     = None

    if request.method == "POST":
        nombre   = request.POST.get("nombre", "").strip()
        correo   = request.POST.get("correo", "").strip()
        password = request.POST.get("password", "").strip()

        if not correo or not password:
            error = "Correo y contraseña son obligatorios"
        elif error_pwd := _error_password(password):
            error = error_pwd
        elif Usuario.objects.filter(correo=correo).exists():
            error = "Ya existe un usuario con ese correo"
        else:
            try:
                porcentaje = Decimal(request.POST.get("porcentaje_ganancia", "0") or "0")
            except Exception:
                porcentaje = Decimal("0")

            inspector = Usuario.objects.create_user(
                correo=correo,
                password=password,
                municipio=municipio,
                es_inspector=True,
                es_conductor=False,
                porcentaje_ganancia=porcentaje,
                periodicidad_rendicion=request.POST.get("periodicidad_rendicion", "semanal"),
            )
            inspector.first_name    = nombre
            inspector.telefono      = request.POST.get("telefono", "").strip()
            inspector.numero_dni    = request.POST.get("numero_dni", "").strip()
            inspector.numero_legajo = request.POST.get("numero_legajo", "").strip()
            inspector.save()
            return redirect("gestionar_inspectores")

    inspectores = Usuario.objects.filter(
        es_inspector=True, municipio=municipio
    ).annotate(
        total_infracciones=Count("infraccion", distinct=True),
        total_verificaciones=Count("verificacioninspector", distinct=True),
        total_cobrado_sum=Sum(
            "movimientocaja__monto",
            filter=Q(movimientocaja__tipo="ingreso"),
        ),
    )

    return render(request, "admin/gestionar_inspectores.html", {
        "inspectores": inspectores,
        "error":       error,
    })


@require_role("admin")
def editar_inspector(request, inspector_id):
    """Edita datos personales y configuración de rendición de un inspector."""
    inspector = get_object_or_404(
        Usuario, id=inspector_id, es_inspector=True, municipio=request.user.municipio
    )

    if request.method == "POST":
        inspector.first_name    = request.POST.get("nombre", "").strip()
        inspector.is_active     = request.POST.get("activo") == "on"
        inspector.telefono      = request.POST.get("telefono", "").strip()
        inspector.numero_dni    = request.POST.get("numero_dni", "").strip()
        inspector.numero_legajo = request.POST.get("numero_legajo", "").strip()

        try:
            inspector.saldo_limite = Decimal(request.POST.get("saldo_limite", "0") or "0")
        except Exception:
            inspector.saldo_limite = 0
        try:
            inspector.porcentaje_ganancia = Decimal(
                request.POST.get("porcentaje_ganancia", "0") or "0"
            )
        except Exception:
            inspector.porcentaje_ganancia = 0

        inspector.periodicidad_rendicion = request.POST.get("periodicidad_rendicion", "semanal")
        inspector.save()
        return redirect("gestionar_inspectores")

    movimientos = MovimientoCaja.objects.filter(
        usuario=inspector
    ).order_by("-creado_en")[:20]

    return render(request, "admin/editar_inspector.html", {
        "inspector":   inspector,
        "movimientos": movimientos,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Gestión de vendedores
# ─────────────────────────────────────────────────────────────────────────────



@require_role("admin")
def historial_vendedor(request, vendedor_id):
    """Historial de movimientos de caja de un vendedor/kiosco.

    Muestra todos los MovimientoCaja del vendedor con filtros opcionales
    por fecha. Incluye totales del período filtrado.
    """
    vendedor  = get_object_or_404(
        Usuario, id=vendedor_id, es_vendedor=True, municipio=request.user.municipio
    )
    movimientos = MovimientoCaja.objects.filter(
        usuario=vendedor
    ).order_by("-creado_en")

    # Filtros opcionales por fecha
    desde = request.GET.get("desde", "").strip()
    hasta = request.GET.get("hasta", "").strip()

    if desde:
        try:
            from datetime import datetime
            movimientos = movimientos.filter(
                creado_en__date__gte=datetime.strptime(desde, "%Y-%m-%d").date()
            )
        except ValueError:
            desde = ""
    if hasta:
        try:
            from datetime import datetime
            movimientos = movimientos.filter(
                creado_en__date__lte=datetime.strptime(hasta, "%Y-%m-%d").date()
            )
        except ValueError:
            hasta = ""

    # Totales del período filtrado
    from django.db.models import Sum
    totales = movimientos.aggregate(
        total_ingresos=Sum("monto", filter=Q(tipo="ingreso")),
        total_egresos=Sum("monto",  filter=Q(tipo="egreso")),
        total_comisiones=Sum("comision_monto"),
    )
    total_ingresos   = totales["total_ingresos"]  or 0
    total_egresos    = totales["total_egresos"]   or 0
    total_comisiones = totales["total_comisiones"] or 0
    neto_municipio   = total_ingresos - total_egresos - total_comisiones

    return render(request, "admin/historial_vendedor.html", {
        "vendedor":        vendedor,
        "movimientos":     movimientos,
        "desde":           desde,
        "hasta":           hasta,
        "total_ingresos":  total_ingresos,
        "total_egresos":   total_egresos,
        "total_comisiones": total_comisiones,
        "neto_municipio":  neto_municipio,
    })

@require_role("admin")
def gestionar_vendedores(request):
    """Lista, crea y configura vendedores (kioscos) del municipio."""
    usuario   = request.user
    municipio = usuario.municipio
    error     = None

    if request.method == "POST":
        nombre   = request.POST.get("nombre", "").strip()
        correo   = request.POST.get("correo", "").strip()
        password = request.POST.get("password", "").strip()

        if not correo or not password:
            error = "Correo y contraseña son obligatorios"
        elif error_pwd := _error_password(password):
            error = error_pwd
        elif Usuario.objects.filter(correo=correo).exists():
            error = "Ya existe un usuario con ese correo"
        else:
            try:
                porcentaje = Decimal(request.POST.get("porcentaje_ganancia", "0") or "0")
            except Exception:
                porcentaje = Decimal("0")

            vendedor = Usuario.objects.create_user(
                correo=correo,
                password=password,
                municipio=municipio,
                es_vendedor=True,
                es_conductor=False,
                porcentaje_ganancia=porcentaje,
                periodicidad_rendicion=request.POST.get("periodicidad_rendicion", "semanal"),
            )
            vendedor.first_name         = nombre
            vendedor.nombre_propietario = request.POST.get("nombre_propietario", "").strip()
            vendedor.documento_cuil     = request.POST.get("documento_cuil", "").strip()
            vendedor.telefono           = request.POST.get("telefono", "").strip()
            vendedor.horario_atencion   = request.POST.get("horario_atencion", "").strip()
            vendedor.save()
            return redirect("gestionar_vendedores")

    vendedores = Usuario.objects.filter(es_vendedor=True, municipio=municipio)
    return render(request, "admin/gestionar_vendedores.html", {
        "vendedores": vendedores,
        "error":      error,
    })


@require_role("admin")
def editar_vendedor(request, vendedor_id):
    """Edita datos y configuración de comisión de un vendedor."""
    vendedor = get_object_or_404(
        Usuario, id=vendedor_id, es_vendedor=True, municipio=request.user.municipio
    )

    if request.method == "POST":
        vendedor.first_name         = request.POST.get("nombre", "").strip()
        vendedor.is_active          = request.POST.get("activo") == "on"
        vendedor.nombre_propietario = request.POST.get("nombre_propietario", "").strip()
        vendedor.documento_cuil     = request.POST.get("documento_cuil", "").strip()
        vendedor.telefono           = request.POST.get("telefono", "").strip()
        vendedor.horario_atencion   = request.POST.get("horario_atencion", "").strip()
        try:
            vendedor.saldo_limite = Decimal(request.POST.get("saldo_limite", "0") or "0")
        except Exception:
            vendedor.saldo_limite = 0
        try:
            vendedor.porcentaje_ganancia = Decimal(
                request.POST.get("porcentaje_ganancia", "0") or "0"
            )
        except Exception:
            vendedor.porcentaje_ganancia = 0
        vendedor.periodicidad_rendicion = request.POST.get("periodicidad_rendicion", "semanal")
        vendedor.save()
        return redirect("gestionar_vendedores")

    return render(request, "admin/editar_vendedor.html", {"vendedor": vendedor})


# ─────────────────────────────────────────────────────────────────────────────
# Gestión de conductores
# ─────────────────────────────────────────────────────────────────────────────



@require_role("admin")
def crear_conductor(request):
    """El admin da de alta un conductor manualmente (registro presencial).

    Una vez creado, redirige al detalle para agregar vehículos y exenciones.
    """
    admin     = request.user
    municipio = admin.municipio
    error     = None

    if request.method == "POST":
        nombre   = request.POST.get("nombre", "").strip().title()
        apellido = request.POST.get("apellido", "").strip().title()
        correo   = request.POST.get("correo", "").strip().lower()
        password = request.POST.get("password", "").strip()

        if not all([nombre, apellido, correo, password]):
            error = "Todos los campos son obligatorios."
        elif error_pwd := _error_password(password):
            error = error_pwd
        elif Usuario.objects.filter(correo=correo).exists():
            error = f"Ya existe un usuario con el correo {correo}."
        else:
            conductor = Usuario.objects.create_user(
                correo=correo,
                password=password,
                first_name=nombre,
                last_name=apellido,
                municipio=municipio,
                es_conductor=True,
                es_admin=False,
                es_inspector=False,
                es_vendedor=False,
            )
            messages.success(
                request,
                f"Conductor {nombre} {apellido} creado. "
                "Podés agregarle vehículo y exención desde acá."
            )
            return redirect("detalle_usuario_admin", usuario_id=conductor.id)

    return render(request, "admin/crear_conductor.html", {
        "error": error,
    })

@require_role("admin")
def gestionar_usuarios(request):
    """
    Lista paginada de conductores del municipio con búsqueda por correo o nombre.
    Paginado a 50 por página: evita traer 400+ conductores + sus vehículos a memoria
    cuando el municipio tiene historial largo. prefetch_related se mantiene para
    que cada página no dispare N+1 al acceder a los vehículos de cada conductor.
    """
    usuario   = request.user
    municipio = usuario.municipio

    q = request.GET.get("q", "").strip()
    qs = Usuario.objects.filter(
        es_conductor=True, municipio=municipio
    ).prefetch_related("vehiculos").order_by("first_name", "last_name")

    if q:
        qs = qs.filter(
            Q(correo__icontains=q) | Q(first_name__icontains=q) | Q(last_name__icontains=q)
        )

    paginator  = Paginator(qs, 50)
    numero_pag = request.GET.get("pagina", 1)
    usuarios   = paginator.get_page(numero_pag)

    return render(request, "admin/gestionar_usuarios.html", {
        "usuarios":  usuarios,
        "paginator": paginator,
        "q":         q,
    })


@require_role("admin")
def detalle_usuario_admin(request, usuario_id):
    """
    Detalle de un conductor: datos, vehículos, infracciones recientes.
    Permite agregar vehículos y editar datos básicos del conductor.
    """
    conductor = get_object_or_404(
        Usuario, id=usuario_id, es_conductor=True, municipio=request.user.municipio
    )
    vehiculos = Vehiculo.objects.filter(vehiculousuario__usuario=conductor)
    accion    = request.POST.get("accion") if request.method == "POST" else None

    if accion == "agregar_vehiculo":
        patente = sanitizar_patente(request.POST.get("patente") or "")
        tipo    = request.POST.get("tipo", "auto")
        if tipo not in ("auto", "moto"):
            tipo = "auto"
        if patente:
            vehiculo, creado = Vehiculo.objects.get_or_create(patente=patente)
            if creado:
                vehiculo.tipo = tipo
                vehiculo.save(update_fields=["tipo"])
            VehiculoUsuario.objects.get_or_create(usuario=conductor, vehiculo=vehiculo)
            messages.success(
                request, f"Vehículo {patente} ({vehiculo.get_tipo_display()}) agregado."
            )
            vehiculos = Vehiculo.objects.filter(vehiculousuario__usuario=conductor)

    elif accion == "editar_datos":
        nombre        = request.POST.get("nombre", "").strip()
        apellido      = request.POST.get("apellido", "").strip()
        telefono      = request.POST.get("telefono", "").strip()
        numero_dni    = request.POST.get("numero_dni", "").strip()
        correo_nuevo  = request.POST.get("correo", "").strip().lower()
        es_verificado = request.POST.get("es_verificado") == "1"

        # Validar correo: no vacío y no duplicado
        if correo_nuevo and correo_nuevo != conductor.correo:
            if Usuario.objects.filter(correo=correo_nuevo).exclude(pk=conductor.pk).exists():
                messages.error(request, f"El correo {correo_nuevo} ya está en uso por otro usuario.")
                return redirect("detalle_usuario_admin", usuario_id=conductor.id)
            conductor.correo = correo_nuevo

        if nombre:
            conductor.first_name = nombre.title()
        if apellido:
            conductor.last_name = apellido.title()
        conductor.telefono      = telefono
        conductor.numero_dni    = numero_dni
        conductor.es_verificado = es_verificado
        conductor.save(update_fields=[
            "correo", "first_name", "last_name", "telefono", "numero_dni", "es_verificado"
        ])
        messages.success(request, "Datos actualizados.")

    elif accion == "cambiar_password":
        nueva_password  = request.POST.get("nueva_password", "").strip()
        confirmar       = request.POST.get("confirmar_password", "").strip()
        if not nueva_password:
            messages.error(request, "La contraseña no puede estar vacía.")
        elif nueva_password != confirmar:
            messages.error(request, "Las contraseñas no coinciden.")
        elif len(nueva_password) < 6:
            messages.error(request, "La contraseña debe tener al menos 6 caracteres.")
        else:
            conductor.set_password(nueva_password)
            conductor.save()
            messages.success(request, f"Contraseña de {conductor.correo} actualizada.")

    # Últimas 5 infracciones (preview)
    infracciones = Infraccion.objects.filter(
        vehiculo__vehiculousuario__usuario=conductor,
        municipio=request.user.municipio,
    ).distinct().order_by("-creado_en")[:5]

    return render(request, "admin/detalle_usuario.html", {
        "conductor":   conductor,
        "vehiculos":   vehiculos,
        "infracciones": infracciones,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Infracciones (vista admin)
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def admin_infracciones(request):
    """
    Lista de infracciones con filtros (patente, inspector, estado, fechas).
    Permite anular o cobrar en efectivo directamente desde la vista.
    """
    usuario   = request.user
    municipio = usuario.municipio

    infracciones = Infraccion.objects.filter(municipio=municipio).select_related(
        "vehiculo", "inspector", "subcuadra"
    ).order_by("-creado_en")

    patente      = sanitizar_patente(request.GET.get("patente", ""))
    inspector_id = request.GET.get("inspector", "").strip()
    estado       = request.GET.get("estado", "").strip()
    fecha_desde  = request.GET.get("fecha_desde", "").strip()
    fecha_hasta  = request.GET.get("fecha_hasta", "").strip()

    if patente:
        infracciones = infracciones.filter(vehiculo__patente__icontains=patente)
    if inspector_id:
        infracciones = infracciones.filter(inspector_id=inspector_id)
    if estado:
        infracciones = infracciones.filter(estado=estado)
    if fecha_desde:
        infracciones = infracciones.filter(creado_en__date__gte=fecha_desde)
    if fecha_hasta:
        infracciones = infracciones.filter(creado_en__date__lte=fecha_hasta)

    if request.method == "POST":
        accion        = request.POST.get("accion")
        infraccion_id = request.POST.get("infraccion_id")

        if accion == "anular" and infraccion_id:
            inf = get_object_or_404(Infraccion, id=infraccion_id, municipio=municipio)
            motivo_anulacion = request.POST.get("motivo_anulacion", "").strip()
            if not motivo_anulacion:
                messages.error(request, "Debés ingresar un motivo para anular la infracción.")
                return redirect(request.get_full_path() + f"#inf-{infraccion_id}")
            if inf.estado == "pendiente":
                inf.estado = "anulada"
                inf.motivo_anulacion = motivo_anulacion
                inf.save(update_fields=["estado", "motivo_anulacion"])
                messages.success(request, f"Infracción #{inf.id} anulada.")

        elif accion == "cobrar" and infraccion_id:
            inf = get_object_or_404(Infraccion, id=infraccion_id, municipio=municipio)
            medio_pago_admin = request.POST.get("medio_pago", "efectivo")
            if medio_pago_admin not in MEDIOS_VALIDOS_COBRO:
                medio_pago_admin = "efectivo"
            try:
                inf = cobrar_infraccion_efectivo(infraccion=inf, cobrador=usuario, medio_pago=medio_pago_admin)
            except ValueError as e:
                messages.warning(request, str(e))
                return redirect(request.get_full_path())
            except Exception as e:
                messages.error(request, f"Error al cobrar: {e}")
                return redirect(request.get_full_path())
            return redirect(reverse("ticket_pago_multa", args=[inf.id]))

        return redirect(request.get_full_path())

    inspectores = Usuario.objects.filter(municipio=municipio, es_inspector=True)

    # Conteo de impagas del municipio (sin filtros de la vista, dato global)
    total_impagas = Infraccion.objects.filter(
        municipio=municipio, estado="pendiente"
    ).count()

    # Paginación: 50 por página reemplaza el slice [:200] que ocultaba infracciones
    # más antiguas al superar el límite. El admin ahora puede navegar el historial completo.
    paginator  = Paginator(infracciones, 50)
    numero_pag = request.GET.get("pagina", 1)
    infracciones_pag = paginator.get_page(numero_pag)

    # Permite abrir el modal de detalle al entrar con ?detalle=ID
    detalle_id = request.GET.get("detalle", "").strip()

    # URL para exportar PDF con los mismos filtros activos en la vista
    from urllib.parse import urlencode
    from django.urls import reverse as _reverse
    params_export = {}
    if fecha_desde:
        params_export["desde"] = fecha_desde
    if fecha_hasta:
        params_export["hasta"] = fecha_hasta
    export_pdf_url = _reverse("pdf_infracciones_juzgado") + ("?" + urlencode(params_export) if params_export else "")

    return render(request, "admin/infracciones.html", {
        "infracciones":  infracciones_pag,
        "paginator":     paginator,
        "inspectores":   inspectores,
        "detalle_id":    detalle_id,
        "total_impagas": total_impagas,
        "export_pdf_url": export_pdf_url,
        "filtros": {
            "patente":     patente,
            "inspector":   inspector_id,
            "estado":      estado,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
        },
    })


@require_role("admin")
def comprobante_infraccion(request, infraccion_id):
    """Vista de impresión para comprobante de pago de infracción."""
    infraccion = get_object_or_404(
        Infraccion, id=infraccion_id, municipio=request.user.municipio
    )
    return render(request, "admin/comprobante_infraccion.html", {
        "infraccion": infraccion,
        "municipio":  request.user.municipio,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Tarifas, horarios y días especiales
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def gestionar_tarifas(request):
    """
    Configura tarifas y parámetros del municipio:
    precio hora (auto/moto), monto infracción, abonos, comisión vendedor, tolerancia multa.
    """
    usuario   = request.user
    municipio = usuario.municipio
    error     = None

    if request.method == "POST":
        def _decimal(campo, minimo=0):
            val = request.POST.get(campo, "0").strip() or "0"
            d   = Decimal(val)
            if d < minimo:
                raise ValueError(f"El campo '{campo}' debe ser >= {minimo}.")
            return d

        def _entero(campo, minimo=0):
            val = request.POST.get(campo, "0").strip() or "0"
            n   = int(val)
            if n < minimo:
                raise ValueError(f"El campo '{campo}' debe ser >= {minimo}.")
            return n

        try:
            precio_auto = _decimal("precio_por_hora", minimo=Decimal("0.01"))
            _val_moto   = request.POST.get("precio_por_hora_moto", "").strip()
            precio_moto = Decimal(_val_moto) if _val_moto else None
            monto_inf   = _decimal("monto_infraccion", minimo=0)
            abono_auto  = _decimal("precio_abono_auto", minimo=0)
            abono_moto  = _decimal("precio_abono_moto", minimo=0)
            comision    = _decimal("comision_vendedor", minimo=0)
            tolerancia  = _entero("tolerancia_multa_minutos", minimo=0)

            Tarifa.objects.update_or_create(
                municipio=municipio,
                defaults={
                    "precio_por_hora":      precio_auto,
                    "precio_por_hora_moto": precio_moto,
                    "monto_infraccion":     monto_inf,
                    "precio_abono_auto":    abono_auto,
                    "precio_abono_moto":    abono_moto,
                }
            )

            municipio.comision_vendedor        = comision
            municipio.tolerancia_multa_minutos = tolerancia
            municipio.save(update_fields=["comision_vendedor", "tolerancia_multa_minutos"])

            messages.success(request, "✅ Tarifas y configuración guardadas correctamente.")
            return redirect("gestionar_tarifas")

        except Exception as e:
            error = f"Error al guardar: {e}"

    tarifa_actual = Tarifa.objects.filter(municipio=municipio).first()
    return render(request, "admin/gestionar_tarifas.html", {
        "tarifa_actual": tarifa_actual,
        "municipio":     municipio,
        "error":         error,
    })


@require_role("admin")
def gestionar_horarios(request):
    """Gestión de horarios semanales de cobro por municipio."""
    municipio = request.user.municipio
    DIAS      = HorarioEstacionamiento.DIAS

    if request.method == "POST":
        for dia_num, _ in DIAS:
            activo      = request.POST.get(f"activo_{dia_num}") == "1"
            hora_inicio = request.POST.get(f"hora_inicio_{dia_num}", "").strip()
            hora_fin    = request.POST.get(f"hora_fin_{dia_num}", "").strip()

            HorarioEstacionamiento.objects.update_or_create(
                municipio=municipio,
                dia_semana=dia_num,
                defaults={
                    "hora_inicio": hora_inicio or "08:00",
                    "hora_fin":    hora_fin    or "15:00",
                    "activo":      activo and bool(hora_inicio) and bool(hora_fin),
                }
            )
        return redirect("gestionar_horarios")

    horarios_existentes = {
        h.dia_semana: h
        for h in HorarioEstacionamiento.objects.filter(municipio=municipio)
    }
    dias_con_horario = [
        (dia_num, dia_label, horarios_existentes.get(dia_num))
        for dia_num, dia_label in DIAS
    ]

    return render(request, "admin/gestionar_horarios.html", {
        "dias_con_horario": dias_con_horario,
    })


@require_role("admin")
def gestionar_dias_especiales(request):
    """Alta, baja y listado de días especiales (feriados, festivos, duelos)."""
    municipio = request.user.municipio

    if request.method == "POST":
        accion = request.POST.get("accion")

        if accion == "agregar":
            fecha        = request.POST.get("fecha", "").strip()
            tipo         = request.POST.get("tipo", "feriado")
            descripcion  = request.POST.get("descripcion", "").strip()
            cobro_activo = request.POST.get("cobro_activo") == "1"
            if fecha and descripcion:
                DiaEspecial.objects.update_or_create(
                    municipio=municipio,
                    fecha=fecha,
                    defaults={
                        "tipo":        tipo,
                        "descripcion": descripcion,
                        "cobro_activo": cobro_activo,
                    }
                )

        elif accion == "eliminar":
            dia_id = request.POST.get("dia_id")
            DiaEspecial.objects.filter(id=dia_id, municipio=municipio).delete()

        return redirect("gestionar_dias_especiales")

    dias = DiaEspecial.objects.filter(municipio=municipio).order_by("fecha")
    return render(request, "admin/gestionar_dias_especiales.html", {
        "dias":  dias,
        "tipos": DiaEspecial.TIPOS,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Rendiciones y cierres de caja
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def admin_rendiciones(request):
    """
    Lista todos los cierres de caja del municipio con filtros.
    Permite filtrar por estado (pendiente/certificado), usuario y rango de fechas.
    """
    municipio = getattr(request.user, "municipio", None)

    filtro      = request.GET.get("filtro", "todos")
    usuario_id  = request.GET.get("usuario_id", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()

    cierres = CierreCaja.objects.filter(
        usuario__municipio=municipio
    ).select_related("usuario", "certificado_por").order_by("-fecha_cierre")

    if filtro == "pendientes":
        cierres = cierres.filter(certificado=False)
    elif filtro == "certificados":
        cierres = cierres.filter(certificado=True)

    if usuario_id:
        cierres = cierres.filter(usuario_id=usuario_id)
    if fecha_desde:
        cierres = cierres.filter(fecha_cierre__date__gte=fecha_desde)
    if fecha_hasta:
        cierres = cierres.filter(fecha_cierre__date__lte=fecha_hasta)

    conteo_pendientes = CierreCaja.objects.filter(
        usuario__municipio=municipio, certificado=False,
    ).count()

    paginator = Paginator(cierres, 20)
    page_obj  = paginator.get_page(request.GET.get("page", 1))

    usuarios_con_cierres = Usuario.objects.filter(
        municipio=municipio,
        cierrecaja__isnull=False,
    ).filter(Q(es_inspector=True) | Q(es_vendedor=True)).distinct().order_by("first_name", "correo")

    # Rendiciones propias del admin a tesorería (para que vea cuáles están pendientes de validación)
    mis_rendiciones = Rendicion.objects.filter(
        admin=request.user
    ).order_by("-fecha_hasta")[:20]

    # Comisiones a vendedores (liquidaciones pendientes/depositadas/certificadas)
    liquidaciones = LiquidacionComision.objects.filter(
        municipio=municipio
    ).select_related("vendedor", "rendicion").order_by("-creado_en")[:50]

    seccion = request.GET.get("seccion", "cierres")  # cierres / rendiciones / comisiones / informes

    # ── Tab Informes ────────────────────────────────────────────────────────
    from .models import DestinatarioInforme
    from datetime import date as date_type
    from django.core.mail import EmailMessage

    destinatarios = DestinatarioInforme.objects.filter(municipio=municipio)
    informe_enviado = False
    informe_error   = None

    if request.method == "POST" and request.POST.get("accion_informe"):
        accion_informe = request.POST.get("accion_informe")

        if accion_informe == "agregar_destinatario":
            nombre = request.POST.get("dest_nombre", "").strip()
            correo = request.POST.get("dest_correo", "").strip()
            if nombre and correo:
                DestinatarioInforme.objects.get_or_create(
                    municipio=municipio, correo=correo,
                    defaults={"nombre": nombre},
                )
            return redirect(request.get_full_path().split("?")[0] + "?seccion=informes")

        elif accion_informe == "quitar_destinatario":
            dest_id = request.POST.get("dest_id")
            DestinatarioInforme.objects.filter(id=dest_id, municipio=municipio).delete()
            return redirect(request.get_full_path().split("?")[0] + "?seccion=informes")

        elif accion_informe == "toggle_destinatario":
            dest_id = request.POST.get("dest_id")
            dest = DestinatarioInforme.objects.filter(id=dest_id, municipio=municipio).first()
            if dest:
                dest.activo = not dest.activo
                dest.save(update_fields=["activo"])
            return redirect(request.get_full_path().split("?")[0] + "?seccion=informes")

        elif accion_informe == "enviar_informe":
            # Período
            inf_desde_str = request.POST.get("inf_desde", "")
            inf_hasta_str = request.POST.get("inf_hasta", "")
            try:
                inf_desde = date_type.fromisoformat(inf_desde_str)
            except ValueError:
                inf_desde = date_type.today().replace(day=1)
            try:
                inf_hasta = date_type.fromisoformat(inf_hasta_str)
            except ValueError:
                inf_hasta = date_type.today()

            # Secciones seleccionadas
            secciones = request.POST.getlist("secciones")

            # Destinatarios seleccionados
            dest_ids = request.POST.getlist("dest_ids")
            correos_dest = list(
                DestinatarioInforme.objects
                .filter(id__in=dest_ids, municipio=municipio)
                .values_list("correo", flat=True)
            )

            if not correos_dest:
                informe_error = "Selecciona al menos un destinatario."
            else:
                # Construir cuerpo del email
                nombre_mun = municipio.nombre if municipio else "Municipio"
                periodo_str = f"{inf_desde.strftime('%d/%m/%Y')} al {inf_hasta.strftime('%d/%m/%Y')}"
                asunto = f"Informe {nombre_mun} — {periodo_str}"

                cuerpo_lineas = [
                    f"Informe del Sistema de Estacionamiento Medido — {nombre_mun}",
                    f"Período: {periodo_str}",
                    "",
                ]

                adjuntos = []

                if "rendiciones" in secciones:
                    total_rendido = Rendicion.objects.filter(
                        admin__municipio=municipio,
                        fecha_desde__date__gte=inf_desde,
                        fecha_hasta__date__lte=inf_hasta,
                    ).aggregate(t=Sum("monto_total"))["t"] or 0
                    cierres_cert = CierreCaja.objects.filter(
                        usuario__municipio=municipio,
                        fecha_cierre__date__gte=inf_desde,
                        fecha_cierre__date__lte=inf_hasta,
                        certificado=True,
                    ).aggregate(t=Sum("monto_municipio"))["t"] or 0
                    cuerpo_lineas += [
                        "=== RESUMEN DE RENDICIONES ===",
                        f"Cierres de caja certificados: ${cierres_cert:,.0f}",
                        f"Rendiciones a tesorería: ${total_rendido:,.0f}",
                        "",
                    ]

                if "vendedores" in secciones:
                    from django.db.models import Sum as DjSum
                    vendedores_recap = (
                        MovimientoCaja.objects
                        .filter(usuario__municipio=municipio, tipo="ingreso",
                                creado_en__date__gte=inf_desde, creado_en__date__lte=inf_hasta)
                        .values("usuario__first_name", "usuario__last_name", "usuario__correo")
                        .annotate(total=Sum("monto"))
                        .order_by("-total")
                    )
                    cuerpo_lineas.append("=== RECAUDACIÓN POR VENDEDOR ===")
                    for v in vendedores_recap:
                        nombre_v = f"{v['usuario__first_name']} {v['usuario__last_name']}".strip() or v["usuario__correo"]
                        cuerpo_lineas.append(f"  {nombre_v}: ${v['total']:,.0f}")
                    cuerpo_lineas.append("")

                if "infracciones" in secciones:
                    cant_imp = Infraccion.objects.filter(
                        municipio=municipio, estado="pendiente",
                        creado_en__date__gte=inf_desde, creado_en__date__lte=inf_hasta,
                    ).count()
                    cant_pag = Infraccion.objects.filter(
                        municipio=municipio, estado="pagada",
                        creado_en__date__gte=inf_desde, creado_en__date__lte=inf_hasta,
                    ).count()
                    cuerpo_lineas += [
                        "=== INFRACCIONES ===",
                        f"Impagas del período: {cant_imp}",
                        f"Cobradas del período: {cant_pag}",
                        "",
                    ]
                    # Adjuntar PDF del juzgado
                    try:
                        pdf_bytes = _generar_pdf_infracciones_juzgado(municipio, inf_desde, inf_hasta)
                        adjuntos.append((
                            f"infracciones_impagas_{inf_desde.strftime('%Y%m%d')}.pdf",
                            pdf_bytes,
                            "application/pdf",
                        ))
                    except Exception as e:
                        cuerpo_lineas.append(f"(No se pudo adjuntar PDF: {e})")

                cuerpo_lineas += [
                    "---",
                    "Este email fue generado automáticamente por el Sistema de Estacionamiento Medido.",
                ]

                try:
                    email = EmailMessage(
                        subject=asunto,
                        body="\n".join(cuerpo_lineas),
                        to=correos_dest,
                    )
                    for nombre_adj, datos_adj, mime_adj in adjuntos:
                        email.attach(nombre_adj, datos_adj, mime_adj)
                    email.send(fail_silently=False)
                    informe_enviado = True
                    messages.success(request, f"Informe enviado a {len(correos_dest)} destinatario(s).")
                except Exception as e:
                    informe_error = f"Error al enviar: {e}"
                    messages.error(request, informe_error)

            return redirect(request.get_full_path().split("?")[0] + "?seccion=informes")

    return render(request, "admin/rendiciones.html", {
        "cierres":            page_obj,
        "filtro":             filtro,
        "conteo_pendientes":  conteo_pendientes,
        "usuario_id":         usuario_id,
        "fecha_desde":        fecha_desde,
        "fecha_hasta":        fecha_hasta,
        "usuarios_con_cierres": usuarios_con_cierres,
        "page_obj":           page_obj,
        "mis_rendiciones":    mis_rendiciones,
        "liquidaciones":      liquidaciones,
        "seccion":            seccion,
        "destinatarios":      destinatarios,
        "informe_enviado":    informe_enviado,
        "informe_error":      informe_error,
    })


@require_role("admin")
def crear_rendicion(request):
    """
    El admin genera una rendición a tesorería seleccionando cierres de caja certificados.

    Los totales se calculan automáticamente desde los CierreCaja seleccionados:
    - total_efectivo    = suma de cierre.total_efectivo
    - total_digital     = suma de cierre.total_transferencia + cierre.total_digital
                          (desde la perspectiva de tesorería, transferencia y digital/card
                           son lo mismo: no es efectivo que el admin manipula físicamente)
    - total_neto        = total_efectivo + total_digital
    - El admin NO puede escribir los montos — solo certifica lo que el sistema calculó.
    """
    from django.db import transaction as db_transaction

    municipio = request.user.municipio

    if request.method == "POST":
        periodo      = request.POST.get("periodo", "").strip()
        notas        = request.POST.get("notas", "").strip()
        cierre_ids   = request.POST.getlist("cierre_ids")
        comprobante  = request.FILES.get("comprobante_archivo")

        if not periodo:
            messages.error(request, "Indicá el período.")
            return redirect("crear_rendicion")

        if not cierre_ids:
            messages.error(request, "Seleccioná al menos un cierre de caja.")
            return redirect("crear_rendicion")

        try:
            cierre_ids_int = [int(pk) for pk in cierre_ids]
        except ValueError:
            messages.error(request, "IDs de cierres inválidos.")
            return redirect("crear_rendicion")

        with db_transaction.atomic():
            # Traer solo cierres del municipio, certificados y aún sin rendir.
            # select_for_update evita race conditions si dos admins operan simultáneamente.
            cierres = CierreCaja.objects.select_for_update().filter(
                id__in=cierre_ids_int,
                usuario__municipio=municipio,
                certificado=True,
                rendicion__isnull=True,
            )

            if cierres.count() != len(cierre_ids_int):
                messages.error(
                    request,
                    "Algunos cierres seleccionados no son válidos (ya rendidos, no certificados o de otro municipio)."
                )
                return redirect("crear_rendicion")

            # Calcular totales desde el desglose de cada cierre
            totales = cierres.aggregate(
                tefectivo     = Sum("total_efectivo"),
                ttransferencia = Sum("total_transferencia"),
                tdigital      = Sum("total_digital"),
            )
            suma_efectivo      = totales["tefectivo"]      or Decimal("0")
            suma_transferencia = totales["ttransferencia"] or Decimal("0")
            suma_digital       = totales["tdigital"]       or Decimal("0")

            # Para tesorería, transferencia + digital (débito/crédito/QR) se agrupan como "digital"
            total_efectivo = suma_efectivo
            total_digital  = suma_transferencia + suma_digital
            total_neto     = total_efectivo + total_digital

            # Fechas del período: min(fecha_cierre) / max(fecha_cierre) de los cierres elegidos
            fechas = cierres.aggregate(
                desde=Min("fecha_cierre"),
                hasta=Max("fecha_cierre"),
            )

            cantidad_cierres = cierres.count()

            rendicion = Rendicion.objects.create(
                municipio           = municipio,
                admin               = request.user,
                periodo             = periodo,
                fecha_desde         = fechas["desde"].date() if fechas["desde"] else date.today(),
                fecha_hasta         = fechas["hasta"].date() if fechas["hasta"] else date.today(),
                total_efectivo      = total_efectivo,
                total_digital       = total_digital,
                total_neto          = total_neto,
                notas_tesorero      = notas,
                comprobante_archivo = comprobante,
            )

            # Vincular cada cierre a esta rendición (auditoría)
            cierres.update(rendicion=rendicion)

        messages.success(
            request,
            f"Rendición creada con {cantidad_cierres} cierre(s). Total neto: ${total_neto:,.2f}"
        )
        return redirect("admin_rendiciones")

    # GET: mostrar cierres disponibles (certificados y sin rendir)
    cierres_pendientes = CierreCaja.objects.filter(
        usuario__municipio=municipio,
        certificado=True,
        rendicion__isnull=True,
    ).select_related("usuario").order_by("fecha_cierre")

    # Pre-calcular totales de todos los cierres disponibles para mostrar en el resumen
    totales_disponibles = cierres_pendientes.aggregate(
        cantidad          = Count("id"),
        suma_efectivo     = Sum("total_efectivo"),
        suma_transferencia = Sum("total_transferencia"),
        suma_digital      = Sum("total_digital"),
        suma_cobrado      = Sum("total_cobrado"),
    )

    return render(request, "admin/crear_rendicion.html", {
        "periodos":            Rendicion.PERIODOS,
        "cierres_pendientes":  cierres_pendientes,
        "totales_disponibles": totales_disponibles,
        "hoy":                 date.today(),
    })


@require_role("admin")
def certificar_cierre(request, cierre_id):
    """El admin certifica (audita) un cierre de caja. Solo acepta POST."""
    cierre = get_object_or_404(
        CierreCaja.objects.select_related("usuario"),
        id=cierre_id,
        usuario__municipio=request.user.municipio,
    )

    if request.method != "POST":
        return redirect("admin_rendiciones")

    if cierre.certificado:
        messages.warning(request, "Este cierre ya estaba certificado.")
        return redirect("admin_rendiciones")

    cierre.certificado    = True
    cierre.certificado_en  = timezone.now()
    cierre.certificado_por = request.user
    cierre.save(update_fields=["certificado", "certificado_en", "certificado_por"])

    messages.success(
        request,
        f"✅ Cierre de {cierre.usuario.correo} del {cierre.fecha_cierre:%d/%m/%Y} certificado."
    )
    return redirect("admin_rendiciones")


# ─────────────────────────────────────────────────────────────────────────────
# Verificaciones de identidad y exenciones
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def gestionar_verificaciones(request):
    """
    Lista solicitudes de verificación filtradas por estado.
    Pendiente muestra tanto identidades como exenciones sin resolver.
    """
    municipio     = getattr(request.user, "municipio", None)
    estado_filtro = request.GET.get("estado", "pendiente")

    solicitudes = SolicitudVerificacion.objects.select_related(
        "usuario", "vehiculo"
    ).filter(usuario__municipio=municipio)

    if estado_filtro == "pendiente":
        solicitudes = solicitudes.filter(
            Q(estado="pendiente") | Q(estado_exencion="pendiente")
        )
    elif estado_filtro in ("aprobada", "rechazada"):
        solicitudes = solicitudes.filter(estado=estado_filtro)

    conteo_pendientes = SolicitudVerificacion.objects.filter(
        estado="pendiente", usuario__municipio=municipio
    ).count()

    subcuadras = Subcuadra.objects.filter(municipio=municipio).order_by("calle", "altura")

    return render(request, "admin/gestionar_verificaciones.html", {
        "solicitudes":       solicitudes,
        "estado_filtro":     estado_filtro,
        "conteo_pendientes": conteo_pendientes,
        "subcuadras":        subcuadras,
        "tipos_exencion":    TIPOS_EXENCION,
    })


@require_role("admin")
def resolver_verificacion(request, solicitud_id):
    """
    El admin resuelve una solicitud de verificación.

    Acciones:
      aprobar           → identidad aprobada, notifica por email y app
      rechazar          → identidad rechazada + notas, notifica por email y app
      aprobar_exencion  → aplica exención global o parcial al vehículo
      rechazar_exencion → rechaza exención + notas_exencion_admin
    """
    solicitud = get_object_or_404(
        SolicitudVerificacion.objects.select_related("usuario", "vehiculo"),
        id=solicitud_id,
        usuario__municipio=request.user.municipio,
    )

    if request.method != "POST":
        return redirect("gestionar_verificaciones")

    accion = request.POST.get("accion")

    # ── Identidad ────────────────────────────────────────────────────────────
    if accion == "aprobar":
        solicitud.estado      = "aprobada"
        solicitud.notas_admin = ""
        solicitud.save(update_fields=["estado", "notas_admin"])

        solicitud.usuario.es_verificado = True
        solicitud.usuario.save(update_fields=["es_verificado"])

        messages.success(request, f"✅ Identidad aprobada: {solicitud.usuario.correo}.")

        _enviar_email_verificacion(
            correo=solicitud.usuario.correo,
            nombre=solicitud.nombre or solicitud.usuario.correo,
            aprobado=True,
        )
        Notificacion.objects.create(
            destinatario=solicitud.usuario,
            mensaje="✅ ¡Tu identidad fue verificada! El municipio confirmó tu cuenta.",
        )

    elif accion == "rechazar":
        notas                 = request.POST.get("notas_admin", "").strip()
        solicitud.estado      = "rechazada"
        solicitud.notas_admin = notas
        solicitud.save(update_fields=["estado", "notas_admin"])

        solicitud.usuario.es_verificado = False
        solicitud.usuario.save(update_fields=["es_verificado"])

        messages.warning(request, f"❌ Identidad rechazada: {solicitud.usuario.correo}.")

        _enviar_email_verificacion(
            correo=solicitud.usuario.correo,
            nombre=solicitud.nombre or solicitud.usuario.correo,
            aprobado=False,
            motivo=notas,
        )
        motivo_txt = f" Motivo: {notas}" if notas else ""
        Notificacion.objects.create(
            destinatario=solicitud.usuario,
            mensaje=f"❌ Tu verificación fue rechazada.{motivo_txt} Podés reenviar tu solicitud.",
        )

    # ── Exención ─────────────────────────────────────────────────────────────
    elif accion == "aprobar_exencion":
        vehiculo = solicitud.vehiculo
        if not vehiculo:
            messages.error(request, "La solicitud no tiene vehículo asociado.")
            return redirect("gestionar_verificaciones")

        tipo_exencion  = request.POST.get("tipo_exencion", "")
        es_global      = request.POST.get("exento_global") == "on"
        subcuadra_ids  = request.POST.getlist("subcuadras")
        notas_exencion = request.POST.get("notas_exencion", "").strip()

        vehiculo.tipo_exencion  = tipo_exencion
        vehiculo.notas_exencion = notas_exencion

        if es_global:
            vehiculo.exento_global  = True
            vehiculo.exento_parcial = False
            vehiculo.subcuadras_exentas.clear()
        else:
            vehiculo.exento_global  = False
            vehiculo.exento_parcial = bool(subcuadra_ids)
            vehiculo.subcuadras_exentas.set(subcuadra_ids)

        vehiculo.save()
        solicitud.estado_exencion = "aprobada"
        solicitud.save(update_fields=["estado_exencion"])

        tipo_label = dict(TIPOS_EXENCION).get(tipo_exencion, tipo_exencion)
        messages.success(
            request,
            f"✅ Exención '{tipo_label}' aplicada a {vehiculo.patente}."
            + (" Global." if es_global else f" {len(subcuadra_ids)} subcuadra(s)."),
        )
        Notificacion.objects.create(
            destinatario=solicitud.usuario,
            mensaje=f"✅ Tu exención fue aprobada para el vehículo {vehiculo.patente}.",
        )

    elif accion == "rechazar_exencion":
        notas_exencion_admin              = request.POST.get("notas_exencion_admin", "").strip()
        solicitud.estado_exencion         = "rechazada"
        solicitud.notas_exencion_admin    = notas_exencion_admin
        solicitud.save(update_fields=["estado_exencion", "notas_exencion_admin"])

        vehiculo_patente = solicitud.vehiculo.patente if solicitud.vehiculo else "(sin vehículo)"
        messages.warning(request, f"❌ Exención rechazada para {vehiculo_patente}.")

        motivo_txt = f" Motivo: {notas_exencion_admin}" if notas_exencion_admin else ""
        Notificacion.objects.create(
            destinatario=solicitud.usuario,
            mensaje=f"❌ Tu solicitud de exención fue rechazada.{motivo_txt}",
        )

    return redirect("gestionar_verificaciones")

# ─────────────────────────────────────────────────────────────────────────────
# Vehículos del municipio
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def admin_vehiculos(request):
    """
    Lista todos los vehículos registrados en el municipio.
    Permite filtrar por patente y tipo.
    """
    municipio = request.user.municipio
    patente   = sanitizar_patente(request.GET.get("patente", ""))
    tipo      = request.GET.get("tipo", "").strip()

    vehiculos = (
        Vehiculo.objects
        .filter(municipio=municipio)
        .prefetch_related("vehiculousuario_set__usuario")
        .order_by("patente")
    )

    if patente:
        vehiculos = vehiculos.filter(patente__icontains=patente)
    if tipo:
        vehiculos = vehiculos.filter(tipo=tipo)

    paginator = Paginator(vehiculos, 50)
    page      = request.GET.get("page", 1)
    vehiculos_pag = paginator.get_page(page)

    return render(request, "admin/vehiculos.html", {
        "vehiculos": vehiculos_pag,
        "filtros": {"patente": patente, "tipo": tipo},
    })


# ─────────────────────────────────────────────────────────────────────────────
# Vehículos del municipio
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def admin_vehiculos(request):
    """
    Lista todos los vehículos registrados en el municipio.
    Permite filtrar por patente y tipo.
    """
    municipio = request.user.municipio
    patente   = sanitizar_patente(request.GET.get("patente", ""))
    tipo      = request.GET.get("tipo", "").strip()

    vehiculos = (
        Vehiculo.objects
        .filter(municipio=municipio)
        .prefetch_related("vehiculousuario_set__usuario")
        .order_by("patente")
    )

    if patente:
        vehiculos = vehiculos.filter(patente__icontains=patente)
    if tipo:
        vehiculos = vehiculos.filter(tipo=tipo)

    paginator = Paginator(vehiculos, 50)
    page      = request.GET.get("page", 1)
    vehiculos_pag = paginator.get_page(page)

    return render(request, "admin/vehiculos.html", {
        "vehiculos": vehiculos_pag,
        "filtros": {"patente": patente, "tipo": tipo},
    })


# ─────────────────────────────────────────────────────────────────────────────
# Historial de estacionamientos del municipio
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def admin_estacionamientos(request):
    """
    Historial de estacionamientos del municipio con filtros básicos.
    Útil para auditoría y para verificar el funcionamiento del sistema.
    """
    municipio   = request.user.municipio
    patente     = sanitizar_patente(request.GET.get("patente", ""))
    estado      = request.GET.get("estado", "").strip()
    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()

    estacionamientos = (
        Estacionamiento.objects
        .filter(subcuadra__municipio=municipio)
        .select_related("vehiculo", "usuario", "subcuadra")
        .order_by("-hora_inicio")
    )

    if patente:
        estacionamientos = estacionamientos.filter(vehiculo__patente__icontains=patente)
    if estado:
        estacionamientos = estacionamientos.filter(estado=estado)
    if fecha_desde:
        estacionamientos = estacionamientos.filter(hora_inicio__date__gte=fecha_desde)
    if fecha_hasta:
        estacionamientos = estacionamientos.filter(hora_inicio__date__lte=fecha_hasta)

    paginator = Paginator(estacionamientos, 50)
    page      = request.GET.get("page", 1)
    estacionamientos_pag = paginator.get_page(page)

    return render(request, "admin/estacionamientos.html", {
        "estacionamientos": estacionamientos_pag,
        "filtros": {
            "patente":     patente,
            "estado":      estado,
            "fecha_desde": fecha_desde,
            "fecha_hasta": fecha_hasta,
        },
    })

# ─────────────────────────────────────────────────────────────────────────────
# PDF infracciones impagas — para presentar en juzgado de faltas
# ─────────────────────────────────────────────────────────────────────────────

def _generar_pdf_infracciones_juzgado(municipio, desde, hasta, infracciones_qs=None):
    """
    Genera un PDF con las infracciones impagas del municipio en el rango de fechas.
    Retorna bytes del PDF listo para HttpResponse o adjunto de email.

    Args:
        municipio: objeto Municipio
        desde / hasta: date objects
        infracciones_qs: queryset ya filtrado (opcional). Si None, filtra impagas del rango.
    """
    import io
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph

    if infracciones_qs is None:
        infracciones_qs = (
            Infraccion.objects
            .filter(municipio=municipio, estado="pendiente",
                    creado_en__date__gte=desde, creado_en__date__lte=hasta)
            .select_related("vehiculo", "inspector", "subcuadra")
            .order_by("creado_en")
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )
    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", parent=estilos["Title"], fontSize=14, alignment=TA_CENTER)
    estilo_sub    = ParagraphStyle("sub",    parent=estilos["Normal"], fontSize=9,
                                   textColor=colors.HexColor("#555555"), alignment=TA_CENTER)
    estilo_pie    = ParagraphStyle("pie",    parent=estilos["Normal"], fontSize=8,
                                   textColor=colors.HexColor("#888888"))

    partes = []

    municipio_nombre = municipio.nombre if municipio else "Municipio"
    generado_en = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    partes.append(Paragraph(f"Infracciones impagas — {municipio_nombre}", estilo_titulo))
    partes.append(Paragraph(
        f"Período: {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')} "
        f"&nbsp;|&nbsp; Generado: {generado_en}",
        estilo_sub,
    ))
    partes.append(Spacer(1, 0.6*cm))

    encabezado = ["#Acta", "Fecha", "Patente", "Inspector", "Subcuadra", "Monto", "Días"]
    filas = [encabezado]

    hoy = timezone.localtime().date()
    monto_total = 0
    for inf in infracciones_qs:
        dias = (hoy - timezone.localtime(inf.creado_en).date()).days
        filas.append([
            str(inf.id),
            timezone.localtime(inf.creado_en).strftime("%d/%m/%Y"),
            inf.vehiculo.patente if inf.vehiculo else "—",
            inf.inspector.nombre_completo() if inf.inspector else "—",
            str(inf.subcuadra) if inf.subcuadra else "—",
            f"${inf.monto:,.0f}",
            str(dias),
        ])
        monto_total += inf.monto

    if len(filas) == 1:
        partes.append(Paragraph("Sin infracciones impagas en el período.", estilos["Normal"]))
    else:
        anchos = [1.4*cm, 2.2*cm, 2.2*cm, 4.0*cm, 3.5*cm, 2.0*cm, 1.4*cm]
        tabla = Table(filas, colWidths=anchos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("ALIGN",         (5, 0), (6, -1), "RIGHT"),
            ("ALIGN",         (0, 0), (0, -1), "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        partes.append(tabla)
        partes.append(Spacer(1, 0.5*cm))
        total_actas = len(filas) - 1
        partes.append(Paragraph(
            f"Total: <b>{total_actas}</b> acta{'s' if total_actas != 1 else ''} "
            f"&nbsp;|&nbsp; Monto total adeudado: <b>${monto_total:,.0f}</b>",
            estilos["Normal"],
        ))

    partes.append(Spacer(1, 1.5*cm))
    partes.append(Paragraph(
        "Documento generado automáticamente por el Sistema de Estacionamiento Medido Municipal.",
        estilo_pie,
    ))

    doc.build(partes)
    buffer.seek(0)
    return buffer.read()


@require_role("admin")
def pdf_infracciones_juzgado(request):
    """
    Descarga un PDF con las infracciones impagas del municipio.

    Parámetros GET:
        frecuencia: "diario" | "semanal" | "mensual" (pre-selecciona rango)
        desde / hasta: YYYY-MM-DD (override manual)
    """
    from datetime import date as date_type, timedelta
    from django.http import HttpResponse

    municipio = request.user.municipio
    hoy = timezone.localtime().date()

    frecuencia = request.GET.get("frecuencia", "mensual")
    desde_str  = request.GET.get("desde", "")
    hasta_str  = request.GET.get("hasta", "")

    try:
        desde = date_type.fromisoformat(desde_str)
    except ValueError:
        if frecuencia == "diario":
            desde = hoy
        elif frecuencia == "semanal":
            desde = hoy - timedelta(days=hoy.weekday())
        else:  # mensual
            desde = hoy.replace(day=1)

    try:
        hasta = date_type.fromisoformat(hasta_str)
    except ValueError:
        hasta = hoy

    pdf_bytes = _generar_pdf_infracciones_juzgado(municipio, desde, hasta)
    nombre = f"infracciones_impagas_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Estadísticas de inspectores
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def estadisticas_inspectores(request):
    """
    Estadísticas de actividad de los inspectores del municipio.

    Sin filtro de inspector → comparativa de todos.
    Con ?inspector_id=X → detalle del inspector seleccionado.
    Filtros de fecha: ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD (default: mes actual).

    TODO (mejora futura, venta): agregar campo Municipio.estadisticas_activo
    y chequear aquí para permitir ocultar la vista desde Django Admin.
    """
    from datetime import date as date_type
    from django.db.models import Count, Q, Sum
    from django.db.models.functions import ExtractHour, TruncDay
    from .models import VerificacionInspector

    municipio = request.user.municipio
    hoy = timezone.localtime().date()

    # ── Filtros ──────────────────────────────────────────────────────────────
    desde_str = request.GET.get("desde", "")
    hasta_str = request.GET.get("hasta", "")
    inspector_id = request.GET.get("inspector_id", "")

    try:
        desde = date_type.fromisoformat(desde_str)
    except ValueError:
        desde = hoy.replace(day=1)  # primer día del mes actual

    try:
        hasta = date_type.fromisoformat(hasta_str)
    except ValueError:
        hasta = hoy

    inspector_sel = None
    if inspector_id:
        try:
            inspector_sel = Usuario.objects.get(
                id=int(inspector_id), municipio=municipio, es_inspector=True
            )
        except (Usuario.DoesNotExist, ValueError):
            pass

    # first_name/last_name son las columnas reales en BD
    # (nombre/apellido son @property Python, no columnas — no se pueden usar en order_by)
    inspectores = Usuario.objects.filter(
        municipio=municipio, es_inspector=True
    ).order_by("first_name", "last_name")

    # ── Queryset base ─────────────────────────────────────────────────────────
    verif_qs = VerificacionInspector.objects.filter(
        inspector__municipio=municipio,
        fecha__date__gte=desde,
        fecha__date__lte=hasta,
    )
    inf_qs = Infraccion.objects.filter(
        municipio=municipio,
        creado_en__date__gte=desde,
        creado_en__date__lte=hasta,
    )

    if inspector_sel:
        verif_qs = verif_qs.filter(inspector=inspector_sel)
        inf_qs = inf_qs.filter(inspector=inspector_sel)

    # ── Comparativa por inspector ─────────────────────────────────────────────
    rango_q = Q(
        verificacioninspector__fecha__date__gte=desde,
        verificacioninspector__fecha__date__lte=hasta,
    )
    rango_inf_q = Q(
        infraccion__creado_en__date__gte=desde,
        infraccion__creado_en__date__lte=hasta,
        infraccion__municipio=municipio,
    )
    comparativa = inspectores.annotate(
        total_verificaciones=Count(
            "verificacioninspector", filter=rango_q
        ),
        total_infracciones=Count(
            "infraccion",
            filter=rango_inf_q & ~Q(infraccion__estado="anulada"),
        ),
        infracciones_anuladas=Count(
            "infraccion",
            filter=rango_inf_q & Q(infraccion__estado="anulada"),
        ),
    )

    # ── Totales del período (solo el inspector seleccionado, o todos) ─────────
    totales = verif_qs.aggregate(total=Count("id"))
    total_verificaciones = totales["total"] or 0
    total_infracciones = inf_qs.exclude(estado="anulada").count()
    total_anuladas = inf_qs.filter(estado="anulada").count()
    tasa_infraccion = (
        round(total_infracciones / total_verificaciones * 100, 1)
        if total_verificaciones else 0
    )

    # ── Distribución por hora del día ─────────────────────────────────────────
    por_hora = list(
        verif_qs.annotate(hora=ExtractHour("fecha"))
        .values("hora")
        .annotate(cantidad=Count("id"))
        .order_by("hora")
    )
    # Rellenar horas sin datos con 0 para el gráfico
    hora_map = {h["hora"]: h["cantidad"] for h in por_hora}
    distribucion_horaria = [
        {"hora": h, "cantidad": hora_map.get(h, 0)}
        for h in range(6, 22)  # 6:00 a 21:00
    ]
    max_hora = max((d["cantidad"] for d in distribucion_horaria), default=1) or 1

    # ── Subcuadras patrulladas ─────────────────────────────────────────────────
    subcuadras_stats = (
        verif_qs
        .values("subcuadra__calle", "subcuadra__altura", "subcuadra__id")
        .annotate(cantidad=Count("id"))
        .order_by("-cantidad")[:10]
    )

    # ── Actividad por día (para sparkline de texto) ───────────────────────────
    por_dia = (
        verif_qs
        .annotate(dia=TruncDay("fecha"))
        .values("dia")
        .annotate(verificaciones=Count("id"))
        .order_by("dia")
    )

    return render(request, "admin/estadisticas_inspectores.html", {
        "inspectores": inspectores,
        "inspector_sel": inspector_sel,
        "desde": desde,
        "hasta": hasta,
        "comparativa": comparativa,
        "total_verificaciones": total_verificaciones,
        "total_infracciones": total_infracciones,
        "total_anuladas": total_anuladas,
        "tasa_infraccion": tasa_infraccion,
        "distribucion_horaria": distribucion_horaria,
        "max_hora": max_hora,
        "subcuadras_stats": subcuadras_stats,
        "por_dia": por_dia,
    })


@require_role("admin")
def estadisticas_inspectores_excel(request):
    """
    Exporta las estadísticas de inspectores al período filtrado como .xlsx.
    Acepta los mismos parámetros GET que estadisticas_inspectores:
    ?desde=YYYY-MM-DD&hasta=YYYY-MM-DD&inspector_id=X
    """
    import io
    from datetime import date as date_type
    from django.db.models import Count, Q
    from django.http import HttpResponse
    from .models import VerificacionInspector
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    municipio = request.user.municipio
    hoy = timezone.localtime().date()

    # ── Filtros (misma lógica que la vista principal) ─────────────────────────
    desde_str = request.GET.get("desde", "")
    hasta_str = request.GET.get("hasta", "")
    inspector_id = request.GET.get("inspector_id", "")

    try:
        desde = date_type.fromisoformat(desde_str)
    except ValueError:
        desde = hoy.replace(day=1)

    try:
        hasta = date_type.fromisoformat(hasta_str)
    except ValueError:
        hasta = hoy

    inspector_sel = None
    if inspector_id:
        try:
            inspector_sel = Usuario.objects.get(
                id=int(inspector_id), municipio=municipio, es_inspector=True
            )
        except (Usuario.DoesNotExist, ValueError):
            pass

    # ── Datos de comparativa ──────────────────────────────────────────────────
    rango_q = Q(
        verificacioninspector__fecha__date__gte=desde,
        verificacioninspector__fecha__date__lte=hasta,
    )
    rango_inf_q = Q(
        infraccion__creado_en__date__gte=desde,
        infraccion__creado_en__date__lte=hasta,
        infraccion__municipio=municipio,
    )
    qs_inspectores = Usuario.objects.filter(
        municipio=municipio, es_inspector=True
    )
    if inspector_sel:
        qs_inspectores = qs_inspectores.filter(pk=inspector_sel.pk)

    comparativa = qs_inspectores.order_by("first_name", "last_name").annotate(
        total_verificaciones=Count("verificacioninspector", filter=rango_q),
        total_infracciones=Count(
            "infraccion",
            filter=rango_inf_q & ~Q(infraccion__estado="anulada"),
        ),
        infracciones_anuladas=Count(
            "infraccion",
            filter=rango_inf_q & Q(infraccion__estado="anulada"),
        ),
    )

    # ── Armar el Excel ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estadísticas inspectores"

    # Estilo de encabezado
    COLOR_HEADER = "1a4d6e"  # azul oscuro
    estilo_header = Font(bold=True, color="FFFFFF")
    relleno_header = PatternFill("solid", fgColor=COLOR_HEADER)
    centrado = Alignment(horizontal="center")

    # ── Título del reporte ────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    celda_titulo = ws["A1"]
    celda_titulo.value = f"Estadísticas de inspectores — {municipio.nombre}"
    celda_titulo.font = Font(bold=True, size=13)
    celda_titulo.alignment = centrado

    ws.merge_cells("A2:F2")
    celda_periodo = ws["A2"]
    celda_periodo.value = f"Período: {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}"
    celda_periodo.alignment = centrado

    ws.append([])  # fila vacía

    # ── Encabezados de la tabla ───────────────────────────────────────────────
    encabezados = [
        "Inspector",
        "Verificaciones",
        "Infracciones",
        "Anuladas",
        "Tasa infracción (%)",
        "Efectividad (%)",
    ]
    ws.append(encabezados)
    fila_header = ws.max_row
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_header, column=col_idx)
        celda.font = estilo_header
        celda.fill = relleno_header
        celda.alignment = centrado

    # ── Filas de datos ────────────────────────────────────────────────────────
    for insp in comparativa:
        nombre_completo = f"{insp.first_name} {insp.last_name}".strip() or insp.correo
        verif = insp.total_verificaciones
        inf   = insp.total_infracciones
        anul  = insp.infracciones_anuladas
        # tasa = infracciones / verificaciones (cuántas verificaciones terminaron en infracción)
        tasa = round(inf / verif * 100, 1) if verif else 0
        # efectividad = (inf - anuladas) / inf (de las que labró, cuántas sobrevivieron)
        efectividad = round((inf - anul) / inf * 100, 1) if inf else 0
        ws.append([nombre_completo, verif, inf, anul, tasa, efectividad])

    # ── Fila de totales ───────────────────────────────────────────────────────
    total_verif = sum(i.total_verificaciones for i in comparativa)
    total_inf   = sum(i.total_infracciones   for i in comparativa)
    total_anul  = sum(i.infracciones_anuladas for i in comparativa)
    tasa_total  = round(total_inf / total_verif * 100, 1) if total_verif else 0

    ws.append([])  # separador
    fila_totales = ["TOTAL", total_verif, total_inf, total_anul, tasa_total, ""]
    ws.append(fila_totales)
    fila_tot_idx = ws.max_row
    for col_idx in range(1, 7):
        celda = ws.cell(row=fila_tot_idx, column=col_idx)
        celda.font = Font(bold=True)

    # ── Ancho de columnas ─────────────────────────────────────────────────────
    anchos = [30, 16, 14, 12, 20, 18]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    # ── Respuesta HTTP como descarga ──────────────────────────────────────────
    nombre_archivo = (
        f"inspectores_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.xlsx"
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


# ─── PDF de rendición ────────────────────────────────────────────────────────

def _generar_pdf_rendicion(rendicion):
    """
    Genera el PDF de una rendición para tesorería.

    Incluye:
    - Encabezado: municipio, período, admin, estado
    - Resumen: efectivo / digital / neto
    - Tabla de detalle: un fila por CierreCaja incluido en la rendición
    - Pie: fecha de generación + notas del tesorero si las hay

    Retorna bytes del PDF listos para HttpResponse.
    """
    import io
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph, HRFlowable

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
    )

    estilos = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle(
        "titulo", parent=estilos["Title"], fontSize=14, alignment=TA_CENTER, spaceAfter=4,
    )
    estilo_sub = ParagraphStyle(
        "sub", parent=estilos["Normal"], fontSize=9,
        textColor=colors.HexColor("#555555"), alignment=TA_CENTER, spaceAfter=2,
    )
    estilo_seccion = ParagraphStyle(
        "seccion", parent=estilos["Normal"], fontSize=10,
        textColor=colors.HexColor("#2c3e50"), fontName="Helvetica-Bold", spaceBefore=10, spaceAfter=4,
    )
    estilo_pie = ParagraphStyle(
        "pie", parent=estilos["Normal"], fontSize=8,
        textColor=colors.HexColor("#888888"), alignment=TA_CENTER,
    )
    estilo_notas = ParagraphStyle(
        "notas", parent=estilos["Normal"], fontSize=9,
        textColor=colors.HexColor("#664d03"),
        backColor=colors.HexColor("#fff3cd"),
        borderPadding=(4, 8, 4, 8),
    )

    municipio_nombre = rendicion.municipio.nombre if rendicion.municipio else "Municipio"
    admin_nombre = rendicion.admin.nombre_completo() if rendicion.admin else "—"
    generado_en = timezone.localtime().strftime("%d/%m/%Y %H:%M")

    # Estado con color de texto (no hay color en PDF inline, usamos texto)
    estado_texto = {
        "pendiente": "Pendiente de validación",
        "validada":  "Validada por tesorería",
        "observada": "Con observaciones",
    }.get(rendicion.estado, rendicion.estado)

    partes = []

    # ── Encabezado ──────────────────────────────────────────────────────────
    partes.append(Paragraph(f"Rendición — {municipio_nombre}", estilo_titulo))
    partes.append(Paragraph(
        f"Período: {rendicion.fecha_desde.strftime('%d/%m/%Y')} al {rendicion.fecha_hasta.strftime('%d/%m/%Y')}"
        f"&nbsp;|&nbsp; Admin: {admin_nombre}"
        f"&nbsp;|&nbsp; Estado: {estado_texto}",
        estilo_sub,
    ))
    partes.append(Paragraph(f"Generado: {generado_en}", estilo_sub))
    partes.append(Spacer(1, 0.4*cm))
    partes.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#2c3e50")))
    partes.append(Spacer(1, 0.4*cm))

    # ── Resumen de totales ───────────────────────────────────────────────────
    partes.append(Paragraph("Resumen de totales", estilo_seccion))
    resumen_data = [
        ["Concepto", "Monto"],
        ["Efectivo", f"${rendicion.total_efectivo:,.2f}"],
        ["Digital (transferencia + débito/crédito/QR)", f"${rendicion.total_digital:,.2f}"],
        ["Total neto a rendir", f"${rendicion.total_neto:,.2f}"],
    ]
    tabla_resumen = Table(resumen_data, colWidths=[12*cm, 4*cm])
    tabla_resumen.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS",(0, 1), (-1, -2), [colors.white, colors.HexColor("#f5f5f5")]),
        # Fila de total neto en negrita con fondo destacado
        ("BACKGROUND",    (0, 3), (-1, 3), colors.HexColor("#e8f5e9")),
        ("FONTNAME",      (0, 3), (-1, 3), "Helvetica-Bold"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
        ("ALIGN",         (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    partes.append(tabla_resumen)
    partes.append(Spacer(1, 0.5*cm))

    # ── Detalle de cierres incluidos ─────────────────────────────────────────
    cierres = list(
        rendicion.cierres.select_related("usuario", "certificado_por").order_by("fecha_cierre")
    )

    if cierres:
        partes.append(Paragraph(f"Cierres de caja incluidos ({len(cierres)})", estilo_seccion))

        encabezado = ["Usuario", "Fecha cierre", "Período", "Efectivo", "Transferencia", "Digital", "Total"]
        filas = [encabezado]

        for cierre in cierres:
            filas.append([
                cierre.usuario.nombre_completo() if cierre.usuario else "—",
                timezone.localtime(cierre.fecha_cierre).strftime("%d/%m/%Y"),
                cierre.get_periodo_display() if cierre.periodo else "—",
                f"${cierre.total_efectivo:,.0f}",
                f"${cierre.total_transferencia:,.0f}",
                f"${cierre.total_digital:,.0f}",
                f"${cierre.total_cobrado:,.0f}",
            ])

        anchos = [4.2*cm, 2.2*cm, 1.8*cm, 2.0*cm, 2.5*cm, 1.8*cm, 2.0*cm]
        tabla_cierres = Table(filas, colWidths=anchos, repeatRows=1)
        tabla_cierres.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 8),
            ("FONTSIZE",      (0, 1), (-1, -1), 7.5),
            ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
            ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("ALIGN",         (3, 0), (-1, -1), "RIGHT"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ]))
        partes.append(tabla_cierres)
    else:
        partes.append(Paragraph("Sin cierres de caja vinculados.", estilos["Normal"]))

    # ── Validación de tesorería ──────────────────────────────────────────────
    if rendicion.tesorero and rendicion.validado_en:
        partes.append(Spacer(1, 0.5*cm))
        partes.append(Paragraph("Validación de tesorería", estilo_seccion))
        validado_en = timezone.localtime(rendicion.validado_en).strftime("%d/%m/%Y %H:%M")
        partes.append(Paragraph(
            f"Validada por: <b>{rendicion.tesorero.nombre_completo()}</b> el {validado_en}",
            estilos["Normal"],
        ))

    if rendicion.notas_tesorero:
        partes.append(Spacer(1, 0.3*cm))
        partes.append(Paragraph(f"Observaciones: {rendicion.notas_tesorero}", estilo_notas))

    # ── Pie ──────────────────────────────────────────────────────────────────
    partes.append(Spacer(1, 1*cm))
    partes.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#cccccc")))
    partes.append(Spacer(1, 0.2*cm))
    partes.append(Paragraph(
        "Documento generado automáticamente por el Sistema de Estacionamiento Medido Municipal.",
        estilo_pie,
    ))

    doc.build(partes)
    buffer.seek(0)
    return buffer.read()


@require_role("admin", "tesorero")
def pdf_rendicion(request, rendicion_id):
    """
    Descarga el PDF de una rendición específica.
    Accesible tanto para el admin que la creó como para tesorería.
    """
    municipio = getattr(request.user, "municipio", None)

    # El admin solo puede ver las rendiciones de su municipio.
    # El tesorero también está restringido por municipio.
    rendicion = get_object_or_404(Rendicion, id=rendicion_id, municipio=municipio)

    pdf_bytes = _generar_pdf_rendicion(rendicion)

    nombre_archivo = (
        f"rendicion_{rendicion.fecha_desde.strftime('%Y%m%d')}"
        f"_{rendicion.fecha_hasta.strftime('%Y%m%d')}.pdf"
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# Gestión de subcuadras y coordenadas GPS
# ─────────────────────────────────────────────────────────────────────────────

@require_role("admin")
def gestionar_subcuadras(request):
    """
    Permite al admin ver, crear, editar y eliminar subcuadras del municipio,
    y asignarles coordenadas GPS haciendo click en un mapa Leaflet/OSM.

    POST accion=guardar_coordenadas: guarda lat/lon para una subcuadra.
    POST accion=limpiar_coordenadas: elimina lat/lon de una subcuadra.
    POST accion=crear:              crea una nueva subcuadra.
    POST accion=eliminar:           elimina una subcuadra sin infracciones.
    """
    municipio = getattr(request.user, "municipio", None)
    if not municipio:
        return redirect("login")

    if request.method == "POST":
        accion = request.POST.get("accion")

        if accion == "guardar_coordenadas":
            subcuadra_id = request.POST.get("subcuadra_id")
            try:
                lat = float(request.POST.get("lat", ""))
                lon = float(request.POST.get("lon", ""))
            except (TypeError, ValueError):
                messages.error(request, "Coordenadas inválidas.")
                return redirect("gestionar_subcuadras")

            sub = get_object_or_404(Subcuadra, id=subcuadra_id, municipio=municipio)
            sub.lat = round(lat, 6)
            sub.lon = round(lon, 6)
            sub.save(update_fields=["lat", "lon"])
            messages.success(request, f"✅ Coordenadas guardadas para {sub}.")

        elif accion == "limpiar_coordenadas":
            sub = get_object_or_404(
                Subcuadra, id=request.POST.get("subcuadra_id"), municipio=municipio
            )
            sub.lat = None
            sub.lon = None
            sub.save(update_fields=["lat", "lon"])
            messages.success(request, f"Coordenadas eliminadas de {sub}.")

        elif accion == "crear":
            calle  = request.POST.get("calle", "").strip()
            altura = request.POST.get("altura", "").strip()
            if not calle or not altura.lstrip("-").isdigit():
                messages.error(request, "Calle y altura son obligatorias.")
            else:
                _, creada = Subcuadra.objects.get_or_create(
                    municipio=municipio,
                    calle=calle,
                    altura=int(altura),
                )
                if creada:
                    messages.success(request, f"✅ Subcuadra '{calle} {altura}' creada.")
                else:
                    messages.warning(request, f"Ya existía la subcuadra '{calle} {altura}'.")

        elif accion == "eliminar":
            sub = get_object_or_404(
                Subcuadra, id=request.POST.get("subcuadra_id"), municipio=municipio
            )
            nombre = str(sub)
            # No eliminar si tiene infracciones, estacionamientos o exenciones asociadas
            en_uso = (
                sub.infraccion_set.exists()
                or sub.estacionamiento_set.exists()
                or sub.vehiculos_exentos_en.exists()
            )
            if en_uso:
                messages.error(
                    request,
                    f"No se puede eliminar '{nombre}' porque tiene registros asociados.",
                )
            else:
                sub.delete()
                messages.success(request, f"Subcuadra '{nombre}' eliminada.")

        return redirect("gestionar_subcuadras")

    # GET: listar subcuadras del municipio
    subcuadras = Subcuadra.objects.filter(municipio=municipio).order_by("calle", "altura")

    # Datos para el mapa: solo las que tienen coordenadas cargadas
    import json as _json
    marcadores = _json.dumps([
        {
            "id":     s.id,
            "nombre": str(s),
            "lat":    float(s.lat),
            "lon":    float(s.lon),
        }
        for s in subcuadras if s.lat is not None and s.lon is not None
    ])

    return render(request, "admin/subcuadras.html", {
        "subcuadras": subcuadras,
        "marcadores": marcadores,
    })


# ─────────────────────────────────────────────────────────────────────────────
# Importación de exenciones desde Excel
# ─────────────────────────────────────────────────────────────────────────────

def _procesar_fila_exencion(fila_num, fila_vals, municipio):
    """
    Procesa una fila del Excel y devuelve un dict con el resultado del análisis.

    Devuelve:
        {
            "num":       int,          # número de fila (para mostrar al usuario)
            "patente":   str,
            "nombre":    str,
            "telefono":  str,
            "direccion": str,          # texto original de la columna Direccion
            "subcuadra": Subcuadra | None,
            "notas":     str,          # texto a guardar en notas_exencion
            "estado":    "nuevo" | "actualizar" | "error",
            "mensaje":   str,
        }
    """
    import re

    # ── Leer celdas ───────────────────────────────────────────────────────────
    def celda(idx):
        """Devuelve el valor de la celda como string limpio (o vacío)."""
        val = fila_vals[idx] if idx < len(fila_vals) else None
        if val is None:
            return ""
        return str(val).strip()

    patente   = re.sub(r"[^A-Z0-9]", "", celda(0).upper())
    nombre    = celda(1)
    direccion = celda(2)
    telefono  = celda(3)
    # columnas Fecha(4), Condicion(5), Vencimiento(6) — las guardamos en notas
    fecha      = celda(4)
    condicion  = celda(5)
    vencimiento = celda(6)

    # ── Validación básica ─────────────────────────────────────────────────────
    if not patente:
        return {
            "num": fila_num, "patente": "", "nombre": nombre,
            "telefono": telefono, "direccion": direccion,
            "subcuadra": None, "notas": "",
            "estado": "error", "mensaje": "Patente vacía — fila ignorada.",
        }

    # ── Buscar subcuadra por nombre de calle (coincidencia parcial) ───────────
    # Se filtra solo dentro del municipio para no mezclar con otros.
    # Si hay más de una coincidencia se toma la primera alfabéticamente.
    subcuadra = None
    aviso_subcuadra = ""
    if direccion:
        candidatas = Subcuadra.objects.filter(
            municipio=municipio,
            calle__icontains=direccion,
        ).order_by("calle", "altura")
        if candidatas.exists():
            subcuadra = candidatas.first()
            if candidatas.count() > 1:
                aviso_subcuadra = f" (múltiples coincidencias para '{direccion}', se tomó la primera)"
        else:
            aviso_subcuadra = f" (sin subcuadra para '{direccion}')"

    # ── Construir notas_exencion ──────────────────────────────────────────────
    partes = []
    if nombre:
        partes.append(f"Nombre: {nombre}")
    if telefono:
        partes.append(f"Tel: {telefono}")
    if fecha:
        partes.append(f"Fecha: {fecha}")
    if direccion:
        partes.append(f"Dirección: {direccion}")
    if condicion:
        partes.append(f"Condición: {condicion}")
    if vencimiento:
        partes.append(f"Vencimiento: {vencimiento}")
    notas = " | ".join(partes)

    # ── Determinar si el vehículo ya existe ───────────────────────────────────
    vehiculo_existente = Vehiculo.objects.filter(patente=patente).first()
    estado  = "actualizar" if vehiculo_existente else "nuevo"
    mensaje = (
        f"Actualiza vehículo existente.{aviso_subcuadra}"
        if vehiculo_existente
        else f"Crea vehículo nuevo.{aviso_subcuadra}"
    )

    return {
        "num":       fila_num,
        "patente":   patente,
        "nombre":    nombre,
        "telefono":  telefono,
        "direccion": direccion,
        "subcuadra": subcuadra,
        "notas":     notas,
        "estado":    estado,
        "mensaje":   mensaje,
    }


def _guardar_fila_exencion(datos_fila, municipio):
    """
    Aplica los datos procesados de una fila al modelo Vehiculo.
    Se llama solo tras la confirmación del admin.

    Reglas:
    - get_or_create por patente.
    - exento_parcial = True, exencion_verificada = False.
    - tipo_exencion = "vecino_frentista".
    - Si hay subcuadra → agregarla a subcuadras_exentas (sin borrar las existentes).
    - notas_exencion: si ya tenía notas, se concatena con " | ".
    """
    patente   = datos_fila["patente"]
    subcuadra = datos_fila["subcuadra"]
    notas     = datos_fila["notas"]

    vehiculo, _ = Vehiculo.objects.get_or_create(
        patente=patente,
        defaults={"tipo": "auto", "municipio": municipio},
    )

    vehiculo.exento_parcial      = True
    vehiculo.exencion_verificada = False
    vehiculo.tipo_exencion       = "vecino_frentista"

    # Concatenar notas sin pisar lo que el admin haya escrito antes
    if notas:
        if vehiculo.notas_exencion:
            vehiculo.notas_exencion = vehiculo.notas_exencion + " | " + notas
        else:
            vehiculo.notas_exencion = notas

    vehiculo.save()

    if subcuadra:
        vehiculo.subcuadras_exentas.add(subcuadra)


@require_role("admin")
def importar_exenciones(request):
    """
    Importa exenciones de vecinos frentistas desde un Excel con columnas:
    Patente | Nombre y Apellido | Direccion | Telefono | Fecha | Condicion | Vencimiento

    Flujo:
    1. GET             → muestra el formulario de carga.
    2. POST preview    → lee el Excel, muestra tabla fila x fila sin guardar.
    3. POST confirmar  → guarda todos los registros en la DB.

    Todos los vehículos importados quedan con exencion_verificada=False
    para que el admin los contacte individualmente y complete los datos faltantes.
    """
    import io
    import json as _json
    try:
        import openpyxl
    except ImportError:
        messages.error(request, "Falta instalar openpyxl: pip install openpyxl")
        return redirect("panel_exenciones")

    usuario   = request.user
    municipio = getattr(usuario, "municipio", None)
    if not municipio:
        return redirect("login")

    accion     = request.POST.get("accion", "")
    resultados = None  # lista de dicts para mostrar en el template

    if request.method == "POST" and accion == "preview":
        archivo = request.FILES.get("archivo")
        if not archivo:
            messages.error(request, "Seleccioná un archivo Excel.")
            return render(request, "admin/importar_exenciones.html", {})

        try:
            wb  = openpyxl.load_workbook(io.BytesIO(archivo.read()), data_only=True)
            ws  = wb.active
            filas = list(ws.iter_rows(min_row=2, values_only=True))  # omite encabezado
        except Exception as exc:
            messages.error(request, f"No se pudo leer el Excel: {exc}")
            return render(request, "admin/importar_exenciones.html", {})

        resultados = []
        for idx, fila in enumerate(filas, start=2):
            # Ignorar filas completamente vacías
            if all(v is None or str(v).strip() == "" for v in fila):
                continue
            datos = _procesar_fila_exencion(idx, fila, municipio)
            resultados.append(datos)

        # Serializar para el formulario de confirmación:
        # no podemos reenviar el archivo, así que guardamos los datos en sesión.
        # Excluimos el objeto Subcuadra (no serializable) y guardamos solo el id.
        datos_sesion = []
        for d in resultados:
            if d["estado"] == "error":
                continue
            datos_sesion.append({
                "patente":     d["patente"],
                "subcuadra_id": d["subcuadra"].id if d["subcuadra"] else None,
                "notas":       d["notas"],
            })
        request.session["importar_exenciones_datos"] = datos_sesion

        return render(request, "admin/importar_exenciones.html", {
            "resultados": resultados,
            "modo": "preview",
        })

    elif request.method == "POST" and accion == "confirmar":
        datos_sesion = request.session.pop("importar_exenciones_datos", None)
        if not datos_sesion:
            messages.error(request, "Sesión expirada. Volvé a cargar el archivo.")
            return render(request, "admin/importar_exenciones.html", {})

        creados    = 0
        actualizados = 0

        for item in datos_sesion:
            patente     = item["patente"]
            subcuadra_id = item.get("subcuadra_id")
            notas        = item.get("notas", "")

            subcuadra = None
            if subcuadra_id:
                try:
                    subcuadra = Subcuadra.objects.get(id=subcuadra_id)
                except Subcuadra.DoesNotExist:
                    pass

            existia = Vehiculo.objects.filter(patente=patente).exists()
            _guardar_fila_exencion(
                {"patente": patente, "subcuadra": subcuadra, "notas": notas},
                municipio,
            )
            if existia:
                actualizados += 1
            else:
                creados += 1

        messages.success(
            request,
            f"✅ Importación completada: {creados} vehículos nuevos, "
            f"{actualizados} actualizados. "
            "Todos quedan pendientes de verificación.",
        )
        return redirect("panel_exenciones")

    # GET
    return render(request, "admin/importar_exenciones.html", {})
